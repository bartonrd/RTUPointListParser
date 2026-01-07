#!/usr/bin/env python3
"""
Extract Point Name column values from PDF files.
This script extracts all values under the "Point Name" column header from every table in each PDF.
"""

import sys
import os
import re
import subprocess
import tempfile
import shutil
import traceback
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None
    print("Note: PyPDF2 not available, will use pdfplumber if available")

try:
    import pdfplumber
except ImportError:
    pdfplumber = None
    print("Note: pdfplumber not available")


# Constants for point name extraction
MAX_NAME_TOKENS = 10  # Maximum tokens to collect for point names
MAX_SMALL_NUMBER_LENGTH = 2  # Maximum length for numbers that can be part of names
STATUS_INDICATORS = ['sh1', 'status']  # Filename patterns indicating Status data
ANALOG_INDICATORS = ['sh2', 'analog']  # Filename patterns indicating Analog data


def check_ocr_tools():
    """Check if OCR tools (tesseract and pdftoppm) are available."""
    try:
        subprocess.run(['tesseract', '--version'], capture_output=True, check=True)
        tesseract_available = True
    except (subprocess.CalledProcessError, FileNotFoundError):
        tesseract_available = False
    
    try:
        # pdftoppm returns exit code 99 for -v, so we don't use check=True
        subprocess.run(['pdftoppm', '-v'], capture_output=True)
        pdftoppm_available = True
    except (subprocess.CalledProcessError, FileNotFoundError):
        pdftoppm_available = False
    
    return tesseract_available, pdftoppm_available


def extract_text_with_ocr(pdf_path):
    """Extract text from image-based PDF using OCR."""
    print(f"  Using OCR to extract text...")
    
    # Create temporary directory
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Convert PDF to images
        output_prefix = os.path.join(temp_dir, 'page')
        subprocess.run(
            ['pdftoppm', '-png', pdf_path, output_prefix],
            check=True,
            capture_output=True
        )
        
        # Get all generated images
        image_files = sorted([f for f in os.listdir(temp_dir) if f.endswith('.png')])
        
        if not image_files:
            print(f"  Warning: No images generated from PDF")
            return ""
        
        # Perform OCR on each image
        all_text = []
        for image_file in image_files:
            image_path = os.path.join(temp_dir, image_file)
            result = subprocess.run(
                ['tesseract', image_path, 'stdout'],
                capture_output=True,
                text=True,
                check=True
            )
            all_text.append(result.stdout)
        
        print(f"  OCR completed on {len(image_files)} page(s)")
        return '\n'.join(all_text)
        
    finally:
        # Clean up temporary directory
        shutil.rmtree(temp_dir, ignore_errors=True)


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF, trying different methods."""
    
    # Try pdfplumber first (usually better for tables)
    if pdfplumber:
        try:
            text_parts = []
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
            
            full_text = '\n'.join(text_parts)
            if full_text.strip():
                return full_text
        except Exception as e:
            print(f"  pdfplumber failed: {e}")
    
    # Try PyPDF2
    if PyPDF2:
        try:
            text_parts = []
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
            
            full_text = '\n'.join(text_parts)
            if full_text.strip():
                return full_text
        except Exception as e:
            print(f"  PyPDF2 failed: {e}")
    
    # If no text extracted, try OCR
    tesseract_available, pdftoppm_available = check_ocr_tools()
    
    if tesseract_available and pdftoppm_available:
        try:
            return extract_text_with_ocr(pdf_path)
        except Exception as e:
            print(f"  OCR failed: {e}")
            return ""
    else:
        print(f"  Warning: OCR tools not available (tesseract: {tesseract_available}, pdftoppm: {pdftoppm_available})")
        return ""


def extract_point_names_from_text(text):
    """
    Extract Point Name values from extracted PDF text.
    This looks for table patterns and extracts the Point Name column.
    """
    point_names = []
    lines = text.split('\n')
    
    # Track if we're in a table with Point Name column
    in_table = False
    point_name_col_index = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Check if this is a header line containing "POINT NAME"
        if 'POINT NAME' in line.upper():
            in_table = True
            # Try to determine column structure
            # The Point Name is typically in column 2 or 3
            continue
        
        # Skip obvious metadata lines
        if any(skip in line for skip in ['PLOT BY:', '.dwg', 'DIAG', 'NOTE', 'COEFFICIENT', 'OFFSET']):
            continue
        
        # Check if this looks like a data row (starts with number followed by | or [ or space)
        # Pattern: NUMBER | POINT_NAME or NUMBER  POINT_NAME
        match = re.match(r'^(\d+)\s*[|\[\s]+(.+)', line)
        if match:
            remainder = match.group(2).strip()
            
            # Split by | to get sections
            sections = remainder.split('|')
            first_section = sections[0].strip()
            
            # Extract point name from first section
            point_name = extract_point_name_from_section(first_section)
            
            # Filter out empty, "Spare", and invalid point names
            if point_name and is_valid_point_name(point_name):
                point_names.append(point_name)
    
    return point_names


def is_valid_point_name(name):
    """Check if a point name is valid (not empty, not "Spare", not just artifacts)."""
    if not name or not name.strip():
        return False
    
    name_upper = name.upper().strip()
    
    # Filter out "SPARE" or variations
    if 'SPARE' in name_upper:
        return False
    
    # Filter out names that are just numbers or single characters
    if len(name.strip()) <= 1:
        return False
    
    # Filter out names that are mostly OCR artifacts
    # Check if it contains at least some letters
    if not any(c.isalpha() for c in name):
        return False
    
    return True


def extract_point_name_from_section(text):
    """
    Extract the point name from a text section.
    Point names typically appear at the start, before control/state information.
    """
    # Split into tokens
    tokens = text.split()
    
    if not tokens:
        return ""
    
    # Collect tokens until we hit state keywords or control markers
    name_tokens = []
    stop_keywords = [
        'CLOSE', 'OPEN', 'NORMAL', 'ALARM', 'AUTO', 'SOLID', 'MANUAL',
        'RK', 'DI', '[or', '[ot', '[pI', '[oI', '[dI'
    ]
    
    for token in tokens:
        # Stop if we hit a keyword
        if any(keyword in token.upper() for keyword in stop_keywords):
            break
        
        # Clean OCR artifacts
        cleaned = clean_ocr_artifacts(token)
        
        if not cleaned or len(cleaned) < 1:
            continue
        
        # Skip if it's just a standalone number (unless following "NO." or similar)
        if cleaned.isdigit() and len(name_tokens) > 0:
            # Allow small numbers as part of name
            if len(cleaned) <= MAX_SMALL_NUMBER_LENGTH:
                name_tokens.append(cleaned)
            break
        
        name_tokens.append(cleaned)
        
        # Limit tokens to avoid including too much
        if len(name_tokens) >= MAX_NAME_TOKENS:
            break
    
    result = ' '.join(name_tokens).strip()
    
    # Clean up multiple spaces
    result = re.sub(r'\s+', ' ', result)
    
    return result


def clean_ocr_artifacts(token):
    """Remove common OCR artifacts from a token."""
    # Remove brackets, pipes, underscores
    cleaned = re.sub(r'[|\[\](){}\_]', '', token)
    
    # Fix common OCR character confusions
    if cleaned.startswith('l') and len(cleaned) > 1 and cleaned[1].isupper():
        cleaned = cleaned[1:]  # Remove leading 'l' confused with 'I'
    
    if cleaned.startswith('/') and len(cleaned) > 1:
        cleaned = cleaned[1:]  # Remove leading slash
    
    return cleaned.strip()


def create_output_excel(point_names_by_sheet, output_path):
    """
    Create an Excel file with Point Name values.
    Structure similar to expected output but with only Point Names.
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets for Status and Analog
    for sheet_name in ['Status', 'Analog']:
        ws = wb.create_sheet(sheet_name)
        
        # Add title
        ws.cell(1, 1, f"CONTRL_D DNP {sheet_name} Point List - Point Names Only")
        ws.cell(1, 1).font = Font(bold=True)
        
        # Add header
        ws.cell(3, 1, "Point Names Extracted from PDF")
        ws.cell(3, 1).font = Font(bold=True)
        
        # Add column header
        ws.cell(5, 1, "POINT NAME")
        ws.cell(5, 1).font = Font(bold=True)
        
        # Add point names
        point_names = point_names_by_sheet.get(sheet_name, [])
        for i, name in enumerate(point_names, start=6):
            ws.cell(i, 1, name)
        
        print(f"  {sheet_name}: {len(point_names)} point names")
    
    wb.save(output_path)


def main():
    """Main function to process PDFs and extract Point Names."""
    
    # Parse arguments
    if len(sys.argv) > 1:
        input_folder = sys.argv[1]
    else:
        input_folder = "ExamplePointlists/Example1/Input"
    
    if len(sys.argv) > 2:
        output_folder = sys.argv[2]
    else:
        output_folder = "ExamplePointlists/Example1/TestOutput"
    
    print("Point Name Extractor")
    print("====================")
    print(f"Input folder: {input_folder}")
    print(f"Output folder: {output_folder}")
    print()
    
    # Validate input folder
    if not os.path.exists(input_folder):
        print(f"Error: Input folder does not exist: {input_folder}")
        return 1
    
    # Create output folder if needed
    os.makedirs(output_folder, exist_ok=True)
    
    # Find all PDF files
    pdf_files = [f for f in os.listdir(input_folder) if f.endswith('.pdf')]
    print(f"Found {len(pdf_files)} PDF file(s) to process")
    print()
    
    if not pdf_files:
        print("No PDF files found in input folder.")
        return 1
    
    # Collect point names by sheet type
    point_names_by_sheet = {
        'Status': [],
        'Analog': []
    }
    
    # Process each PDF
    for pdf_file in sorted(pdf_files):
        pdf_path = os.path.join(input_folder, pdf_file)
        print(f"Processing: {pdf_file}")
        
        try:
            # Extract text
            text = extract_text_from_pdf(pdf_path)
            
            if not text.strip():
                print(f"  Warning: No text extracted from PDF")
                continue
            
            # Extract point names
            point_names = extract_point_names_from_text(text)
            
            # Determine sheet type from filename
            filename_lower = pdf_file.lower()
            if any(indicator in filename_lower for indicator in STATUS_INDICATORS):
                sheet_type = 'Status'
            elif any(indicator in filename_lower for indicator in ANALOG_INDICATORS):
                sheet_type = 'Analog'
            else:
                # Default to Status
                sheet_type = 'Status'
            
            point_names_by_sheet[sheet_type].extend(point_names)
            print(f"  Extracted {len(point_names)} point names ({sheet_type})")
            
        except Exception as e:
            print(f"  Error processing {pdf_file}: {e}")
            traceback.print_exc()
    
    print()
    print("Creating output Excel file...")
    
    # Generate output
    output_file = "Point_Names_Extracted.xlsx"
    output_path = os.path.join(output_folder, output_file)
    create_output_excel(point_names_by_sheet, output_path)
    
    print(f"Output saved to: {output_path}")
    print()
    print("Extraction complete.")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
